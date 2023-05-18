from bs4 import BeautifulSoup
import requests
import os
import subprocess
import tkinter as tk
import pandas as pd
from openpyxl import load_workbook
import openpyxl

URL = 'https://www.rbc.ru/'

def get_html(url):
    # Получаем HTML-код страницы по URL
    response = requests.get(url)
    return response.content

def parse_data(html):
    # Извлекаем заголовки и описания новостей из HTML-кода
    soup = BeautifulSoup(html, features='lxml')
    news = soup.find_all('a', class_='main__link') + soup.find_all('a', class_='news-feed__item')
    titles = []
    for item in news:
        title = item.find('span', {'class': lambda x: x and 'item__title' in x})
        if title is not None:
            # Если заголовок найден, сохраняем его текст
            title_text = title.text
            titles.append(title_text)
    return titles

def save_data(titles):
    # Сохраняем заголовки новостей в файл xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Title"])
    for title in titles:
        ws.append([title])
    wb.save("news.xlsx")

def main():
    def check_internet_connection():
        # Функция для проверки наличия интернет-соединения
        try:
            requests.get("https://www.yandex.ru", timeout=3)
            return True
        except requests.ConnectionError:
            return False
    html = get_html(URL)
    titles = parse_data(html)
    save_data(titles)
    # Открываем файл с новостями в блокноте
    # if os.name == 'nt':  # для Windows
    #    os.startfile('news.xlsx')
    # elif os.name == 'posix':  # для MacOS и Linux
    #    subprocess.call(('open', 'news.xlsx'))

class NewsApp:
    def __init__(self, news_file_path):
        wb = load_workbook(news_file_path)
        sheet = wb.active
        data = sheet.values
        columns = next(data)
        self.news_df = pd.DataFrame(data, columns=columns)

        self.root = tk.Tk()
        self.root.title("Новости")

        # Устанавливаем размер окна и разрешение экрана
        window_width = 1024
        window_height = 720
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x_coordinate = int((screen_width - window_width) / 2)
        y_coordinate = int((screen_height - window_height) / 2)
        self.root.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")

        self.news_frame = tk.Frame(self.root)
        self.news_frame.pack()

        self.scrollbar = tk.Scrollbar(self.news_frame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.news_text = tk.Text(self.news_frame, yscrollcommand=self.scrollbar.set)
        self.news_text.pack(side=tk.LEFT, fill=tk.BOTH)

        self.news_text.config(width=window_width // 10, height=window_height // 20)  # Увеличиваем размер текстовой области

        self.news_text.insert(tk.END, '-' * (window_width // 10) + '\n')  # Сплошная полоса
        titles_str = '\n'.join(['-' * 100 + '\n' + title for title in self.news_df['Title'].tolist()])  # Преобразование списка заголовков в одну строку
        self.news_text.insert(tk.END, titles_str)
        self.news_text.insert(tk.END, '\n' + '-' * (window_width // 10))  # Сплошная полоса

        self.scrollbar.config(command=self.news_text.yview)

        self.close_button = tk.Button(self.root, text="Закрыть", command=self.root.quit)
        self.close_button.pack()

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    main()
    news_file_path = "news.xlsx"
    app = NewsApp(news_file_path)
    app.run()
